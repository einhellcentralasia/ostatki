#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Read-only SharePoint folder sync:
- Lists all .xlsx files in a SharePoint folder (optionally recursive)
- For each file, loads Excel Table (SP_TABLE_NAME, default "Table1") fully in memory
- Extracts only SKU and Qty columns (robust header matching)
- Aggregates Qty by SKU across all files
- Saves outputs to repo ./data/:
    - sold_to_clients.parquet
    - sold_to_clients.csv
    - sold_to_clients.json
- Prints runtime + processing stats

Safety (poka-yoke):
- NO write calls to SharePoint (only GET)
- Skips temp files like "~$..."
- Handles Graph pagination + throttling retries
- Robust path resolution (auto-strips "Shared Documents/" prefix + URL encodes Cyrillic/commas)
- If Table metadata isn't visible, fallback scans for headers SKU+Qty
- Never crashes if nothing parsed: still writes empty outputs with correct schema
"""

from __future__ import annotations

import io
import json
import os
import sys
import time
from dataclasses import dataclass
from typing import Dict, Iterable, List, Optional, Tuple

import pandas as pd
import requests
from openpyxl import load_workbook
from urllib.parse import quote


# -------------------------
# Error handling at the top
# -------------------------

def die(msg: str, code: int = 2) -> None:
    print(f"❌ {msg}", file=sys.stderr)
    raise SystemExit(code)

def env(name: str, default: Optional[str] = None, required: bool = True) -> str:
    val = os.getenv(name, default)
    if required and (val is None or str(val).strip() == ""):
        die(f"Missing required env var: {name}")
    return str(val)

def now_ts() -> str:
    return time.strftime("%Y-%m-%d %H:%M:%S")


# -------------------------
# Microsoft Graph helpers
# -------------------------

GRAPH = "https://graph.microsoft.com/v1.0"

@dataclass
class GraphCtx:
    token: str
    session: requests.Session

def new_session() -> requests.Session:
    s = requests.Session()
    s.headers.update({"Accept": "application/json"})
    return s

def request_raw(ctx: GraphCtx, method: str, url: str, *, params=None, timeout: int = 60) -> requests.Response:
    max_tries = 8
    backoff = 1.0
    headers = {"Authorization": f"Bearer {ctx.token}", "Accept": "application/json"}

    for attempt in range(1, max_tries + 1):
        resp = ctx.session.request(method, url, headers=headers, params=params, timeout=timeout)

        if resp.status_code in (429, 500, 502, 503, 504):
            retry_after = resp.headers.get("Retry-After")
            try:
                sleep_s = float(retry_after) if retry_after else backoff
            except ValueError:
                sleep_s = backoff
            print(f"⚠️ {now_ts()} Graph {resp.status_code} on {url} (attempt {attempt}/{max_tries}), sleep {sleep_s:.1f}s")
            time.sleep(sleep_s)
            backoff = min(backoff * 1.8, 20.0)
            continue

        return resp

    die(f"Graph failed after retries on {url}")
    return resp  # unreachable

def request_json_ok(ctx: GraphCtx, method: str, url: str, *, params=None, expected=(200,)) -> dict:
    resp = request_raw(ctx, method, url, params=params, timeout=60)
    if resp.status_code not in expected:
        body = resp.text[:2000]
        die(f"Graph error {resp.status_code} on {url}\nResponse: {body}")
    return resp.json()

def request_bytes(ctx: GraphCtx, url: str) -> bytes:
    max_tries = 8
    backoff = 1.0
    headers = {"Authorization": f"Bearer {ctx.token}"}

    for attempt in range(1, max_tries + 1):
        resp = ctx.session.get(url, headers=headers, stream=True, timeout=120)
        if resp.status_code == 200:
            return resp.content

        if resp.status_code in (429, 500, 502, 503, 504):
            retry_after = resp.headers.get("Retry-After")
            try:
                sleep_s = float(retry_after) if retry_after else backoff
            except ValueError:
                sleep_s = backoff
            print(f"⚠️ {now_ts()} Download {resp.status_code} (attempt {attempt}/{max_tries}), sleep {sleep_s:.1f}s")
            time.sleep(sleep_s)
            backoff = min(backoff * 1.8, 20.0)
            continue

        die(f"Download error {resp.status_code} for {url}: {resp.text[:1000]}")

    die(f"Download failed after retries: {url}")
    return b""

def get_app_token(tenant_id: str, client_id: str, client_secret: str) -> str:
    token_url = f"https://login.microsoftonline.com/{tenant_id}/oauth2/v2.0/token"
    data = {
        "client_id": client_id,
        "client_secret": client_secret,
        "grant_type": "client_credentials",
        "scope": "https://graph.microsoft.com/.default",
    }
    resp = requests.post(token_url, data=data, timeout=60)
    if resp.status_code != 200:
        die(f"Token request failed {resp.status_code}: {resp.text[:2000]}")
    js = resp.json()
    tok = js.get("access_token")
    if not tok:
        die("Token response missing access_token")
    return tok

def graph_get_site_id(ctx: GraphCtx, hostname: str, site_path: str) -> str:
    url = f"{GRAPH}/sites/{hostname}:{site_path}"
    js = request_json_ok(ctx, "GET", url, expected=(200,))
    site_id = js.get("id")
    if not site_id:
        die("Could not resolve site id (missing id). Check SP_SITE_HOSTNAME/SP_SITE_PATH.")
    return site_id

def graph_get_drive_id(ctx: GraphCtx, site_id: str) -> str:
    url = f"{GRAPH}/sites/{site_id}/drive"
    js = request_json_ok(ctx, "GET", url, expected=(200,))
    drive_id = js.get("id")
    if not drive_id:
        die("Could not resolve drive id for the site.")
    return drive_id

def normalize_sp_path(p: str) -> str:
    p2 = p.strip().replace("\\", "/").lstrip("/")
    while "//" in p2:
        p2 = p2.replace("//", "/")

    low = p2.lower()
    for prefix in ("shared documents/", "documents/"):
        if low.startswith(prefix):
            p2 = p2[len(prefix):]
            break
    return p2

def try_get_item_id_by_path(ctx: GraphCtx, drive_id: str, path: str) -> Optional[str]:
    path_clean = normalize_sp_path(path)
    path_enc = quote(path_clean, safe="/")
    url = f"{GRAPH}/drives/{drive_id}/root:/{path_enc}"
    resp = request_raw(ctx, "GET", url, timeout=60)
    if resp.status_code == 200:
        return resp.json().get("id")
    if resp.status_code == 404:
        return None
    die(f"Unexpected status {resp.status_code} resolving folder path.\nURL: {url}\nBody: {resp.text[:2000]}")
    return None

def graph_list_root_children_names(ctx: GraphCtx, drive_id: str, limit: int = 60) -> List[str]:
    url = f"{GRAPH}/drives/{drive_id}/root/children"
    params = {"$top": str(limit)}
    js = request_json_ok(ctx, "GET", url, params=params, expected=(200,))
    return [it.get("name", "") for it in js.get("value", [])]

def graph_list_children(ctx: GraphCtx, drive_id: str, folder_item_id: str) -> Iterable[dict]:
    url = f"{GRAPH}/drives/{drive_id}/items/{folder_item_id}/children"
    params = {"$top": "200"}
    while True:
        js = request_json_ok(ctx, "GET", url, params=params, expected=(200,))
        for it in js.get("value", []):
            yield it
        next_link = js.get("@odata.nextLink")
        if not next_link:
            break
        url = next_link
        params = None

def graph_walk_files(ctx: GraphCtx, drive_id: str, folder_item_id: str, recursive: bool) -> List[dict]:
    out = []
    stack = [folder_item_id]
    while stack:
        fid = stack.pop()
        for it in graph_list_children(ctx, drive_id, fid):
            name = it.get("name", "")
            if name.startswith("~$"):
                continue

            is_folder = "folder" in it
            is_file = "file" in it

            if is_folder and recursive:
                child_id = it.get("id")
                if child_id:
                    stack.append(child_id)
                continue

            if is_file:
                out.append(it)
    return out


# -------------------------
# Excel parsing helpers
# -------------------------

def norm(s: str) -> str:
    return "".join(str(s).strip().lower().split())

SKU_HEADERS = {"sku", "артикул", "арт", "item", "code"}
QTY_HEADERS = {"qty", "qt", "quantity", "кол-во", "количество", "count", "pcs"}

def find_col_indices(headers: List[str]) -> Tuple[Optional[int], Optional[int]]:
    sku_idx = None
    qty_idx = None
    for i, h in enumerate(headers):
        h2 = norm(h)
        if sku_idx is None and h2 in SKU_HEADERS:
            sku_idx = i
        if qty_idx is None and h2 in QTY_HEADERS:
            qty_idx = i
    return sku_idx, qty_idx

def iter_ws_tables(ws):
    t = getattr(ws, "tables", None)
    if not t:
        return
    try:
        # TableList behaves like a dict
        for name in list(t.keys()):
            yield name, t[name]
    except Exception:
        # Fallback: assume iterable of table objects
        try:
            for tbl in t:
                yield getattr(tbl, "name", None), tbl
        except Exception:
            return

def extract_df_from_range(ws, ref: str) -> pd.DataFrame:
    cells = ws[ref]
    rows = [[c.value for c in row] for row in cells]
    if not rows or len(rows) < 2:
        return pd.DataFrame(columns=["SKU", "Qty"])
    headers = [str(x) if x is not None else "" for x in rows[0]]
    sku_idx, qty_idx = find_col_indices(headers)
    if sku_idx is None or qty_idx is None:
        return pd.DataFrame()  # signal "bad headers"
    skus, qtys = [], []
    for r in rows[1:]:
        if sku_idx >= len(r) or qty_idx >= len(r):
            continue
        sku = r[sku_idx]
        qty = r[qty_idx]
        if sku is None or str(sku).strip() == "":
            continue
        try:
            q = float(qty) if qty is not None and str(qty).strip() != "" else 0.0
        except Exception:
            q = 0.0
        skus.append(str(sku).strip())
        qtys.append(q)
    return pd.DataFrame({"SKU": skus, "Qty": qtys})

def fallback_scan_headers(ws, max_rows: int = 5000, max_cols: int = 80) -> Optional[pd.DataFrame]:
    # Find header row containing SKU and Qty (anywhere), then read down until blank block.
    mr = min(ws.max_row or 0, max_rows)
    mc = min(ws.max_column or 0, max_cols)
    if mr <= 0 or mc <= 0:
        return None

    header_row = None
    sku_col = None
    qty_col = None

    for r in range(1, mr + 1):
        values = []
        for c in range(1, mc + 1):
            v = ws.cell(row=r, column=c).value
            values.append("" if v is None else str(v))
        sku_idx, qty_idx = find_col_indices(values)
        if sku_idx is not None and qty_idx is not None:
            header_row = r
            sku_col = sku_idx + 1
            qty_col = qty_idx + 1
            break

    if header_row is None:
        return None

    skus, qtys = [], []
    blank_streak = 0
    for r in range(header_row + 1, mr + 1):
        sku = ws.cell(row=r, column=sku_col).value
        qty = ws.cell(row=r, column=qty_col).value

        is_blank = (sku is None or str(sku).strip() == "") and (qty is None or str(qty).strip() == "")
        if is_blank:
            blank_streak += 1
            if blank_streak >= 3:
                break
            continue
        blank_streak = 0

        if sku is None or str(sku).strip() == "":
            continue
        try:
            q = float(qty) if qty is not None and str(qty).strip() != "" else 0.0
        except Exception:
            q = 0.0

        skus.append(str(sku).strip())
        qtys.append(q)

    return pd.DataFrame({"SKU": skus, "Qty": qtys})

def read_sku_qty_from_xlsx_bytes(xlsx_bytes: bytes, table_name: str, debug_tables: bool = False) -> Tuple[Optional[pd.DataFrame], str]:
    """
    Returns (df, reason).
    df=None means hard failure.
    df empty with columns means ok but no rows.
    reason used for logs.
    """
    # IMPORTANT FIX: read_only=False so table metadata is available reliably
    wb = load_workbook(filename=io.BytesIO(xlsx_bytes), data_only=True, read_only=False)

    # 1) Try real Excel table by name (case-insensitive)
    target = table_name.strip().lower()
    for ws in wb.worksheets:
        found = []
        for tname, tbl in iter_ws_tables(ws) or []:
            if tname:
                found.append(tname)
        if debug_tables and found:
            print(f"   🧩 Sheet '{ws.title}' tables: {found}")

        for tname, tbl in iter_ws_tables(ws) or []:
            if not tname:
                continue
            if tname.strip().lower() == target:
                ref = getattr(tbl, "ref", None)
                if not ref:
                    return None, "Table found but missing ref"
                df = extract_df_from_range(ws, ref)
                if df.empty and list(df.columns) != ["SKU", "Qty"]:
                    # headers mismatch
                    break
                return df, "OK(table)"

    # 2) Fallback scan headers (handles weird table metadata)
    for ws in wb.worksheets:
        df2 = fallback_scan_headers(ws)
        if df2 is not None:
            return df2, "OK(fallback-scan)"

    return None, "Table not found + fallback scan failed"


# -------------------------
# Main
# -------------------------

def main() -> None:
    t0 = time.perf_counter()

    tenant_id = env("TENANT_ID")
    client_id = env("CLIENT_ID")
    client_secret = env("CLIENT_SECRET")

    sp_site_hostname = env("SP_SITE_HOSTNAME")
    sp_site_path = env("SP_SITE_PATH")
    sp_xlsx_path_raw = env("SP_XLSX_PATH")
    sp_table_name = env("SP_TABLE_NAME", default="Table1", required=False)

    recursive = env("SP_RECURSIVE", default="false", required=False).strip().lower() in ("1", "true", "yes", "y")
    debug_tables = env("DEBUG_TABLES", default="false", required=False).strip().lower() in ("1", "true", "yes", "y")

    out_dir = "data"
    base_name = "sold_to_clients"

    print(f"🟢 {now_ts()} Start. Recursive={recursive} DebugTables={debug_tables}")
    print(f"   Site: {sp_site_hostname}{sp_site_path}")
    print(f"   Folder (raw): {sp_xlsx_path_raw}")
    print(f"   Folder (norm): {normalize_sp_path(sp_xlsx_path_raw)}")
    print(f"   Table: {sp_table_name}")

    token = get_app_token(tenant_id, client_id, client_secret)
    ctx = GraphCtx(token=token, session=new_session())

    site_id = graph_get_site_id(ctx, sp_site_hostname, sp_site_path)
    drive_id = graph_get_drive_id(ctx, site_id)

    folder_item_id = try_get_item_id_by_path(ctx, drive_id, sp_xlsx_path_raw)
    if not folder_item_id:
        root_names = graph_list_root_children_names(ctx, drive_id, limit=60)
        die(
            "Folder not found by path (Graph 404 itemNotFound).\n"
            f"Tried normalized path: {normalize_sp_path(sp_xlsx_path_raw)}\n"
            "Top folders at drive root (for debugging):\n"
            + "\n".join([f" - {n}" for n in root_names])
        )

    items = graph_walk_files(ctx, drive_id, folder_item_id, recursive=recursive)
    xlsx_items = [it for it in items if it.get("name", "").lower().endswith(".xlsx")]

    print(f"📦 Found: total_files={len(items)}, xlsx={len(xlsx_items)}")

    agg: Dict[str, float] = {}
    processed = 0
    skipped = 0
    used_fallback = 0

    # For poka-yoke: limit verbose per-file logs
    max_warns = 50
    warns = 0

    for it in xlsx_items:
        name = it.get("name", "")
        item_id = it.get("id")
        if not item_id:
            skipped += 1
            continue

        content_url = f"{GRAPH}/drives/{drive_id}/items/{item_id}/content"
        try:
            b = request_bytes(ctx, content_url)
        except Exception as e:
            if warns < max_warns:
                print(f"⚠️ Skip '{name}': download failed: {e}")
                warns += 1
            skipped += 1
            continue

        try:
            df, reason = read_sku_qty_from_xlsx_bytes(b, sp_table_name, debug_tables=False)
        except Exception as e:
            if warns < max_warns:
                print(f"⚠️ Skip '{name}': parse failed: {e}")
                warns += 1
            skipped += 1
            continue

        if df is None:
            if warns < max_warns:
                print(f"⚠️ Skip '{name}': {reason}")
                warns += 1
            skipped += 1
            continue

        if reason.endswith("(fallback-scan)"):
            used_fallback += 1

        if not df.empty:
            for sku, qty in zip(df["SKU"].tolist(), df["Qty"].tolist()):
                agg[sku] = agg.get(sku, 0.0) + float(qty)

        processed += 1

    # Poka-yoke: always produce a dataframe with correct schema
    if agg:
        result = pd.DataFrame([{"SKU": sku, "Qty": agg[sku]} for sku in agg.keys()])
        result = result.sort_values(["SKU"], kind="stable")
        result["Qty"] = result["Qty"].apply(lambda x: int(x) if float(x).is_integer() else float(x))
    else:
        result = pd.DataFrame(columns=["SKU", "Qty"])

    os.makedirs(out_dir, exist_ok=True)

    parquet_path = os.path.join(out_dir, f"{base_name}.parquet")
    csv_path = os.path.join(out_dir, f"{base_name}.csv")
    json_path = os.path.join(out_dir, f"{base_name}.json")

    result.to_parquet(parquet_path, index=False)
    result.to_csv(csv_path, index=False, encoding="utf-8")
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(result.to_dict(orient="records"), f, ensure_ascii=False, indent=2)

    dur = time.perf_counter() - t0
    print("✅ Done.")
    print(f"🧾 Processed xlsx: {processed} | Skipped: {skipped} | Used fallback: {used_fallback}")
    print(f"🧮 Unique SKUs: {len(result)}")
    print(f"⏱️ Runtime: {dur:.2f} seconds")
    print(f"📁 Outputs: {parquet_path}, {csv_path}, {json_path}")

    # Optional deep debug for 1 file if you want: set DEBUG_TABLES=true and rerun
    if debug_tables:
        print("🔎 DEBUG_TABLES=true: rerun will print detected table names per sheet for troubleshooting.")


if __name__ == "__main__":
    main()
